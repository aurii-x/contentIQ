#!/usr/bin/env python3
"""
parse_linkedin_export.py

Reads a LinkedIn "Content Analytics" export (.xlsx) and merges post-level
performance into data/master.csv, matching the ContentIQ schema:

    date, post_url, post_topic, content_type, pillar, impressions,
    reactions, comments, shares, total_engagements, engagement_rate,
    new_followers

Notes on what LinkedIn's export actually gives us:
  - The TOP POSTS sheet has two side-by-side tables: one ranked by
    Engagements, one ranked by Impressions. They are NOT row-aligned to
    the same post, they must be merged by Post URL.
  - LinkedIn only gives a single "Engagements" total per post, not a
    breakdown of reactions/comments/shares. Those three columns are left
    blank for manual entry if you want that level of detail later.
  - content_type and pillar are not in the export at all, tag those by
    hand after running this script (or fill them in before re-running,
    your manual edits are preserved on re-run).
  - New followers is a daily total, not a per-post number. It's attached
    to the first post published that day as an approximation, later
    posts on the same day get blank rather than double-counting.

Usage:
    python3 parse_linkedin_export.py path/to/export.xlsx
"""

import csv
import sys
import os
import re
import unicodedata
from datetime import datetime
from urllib.parse import unquote
from pathlib import Path

import openpyxl

MASTER_PATH = Path("src/_data/master.csv")
DAILY_PATH = Path("src/_data/daily_totals.csv")
DISCOVERY_LOG_PATH = Path("src/_data/discovery_log.csv")
SNAPSHOT_PATH = Path("src/_data/post_snapshots.csv")

FIELDNAMES = [
    "date", "post_url", "post_topic", "content_type", "pillar",
    "impressions", "reactions", "comments", "shares",
    "total_engagements", "engagement_rate", "new_followers",
]

DAILY_FIELDNAMES = [
    "date", "impressions", "engagements", "engagement_rate", "new_followers",
]

DISCOVERY_FIELDNAMES = [
    "pulled_at", "range_start", "range_end", "total_impressions", "total_members_reached",
]

SNAPSHOT_FIELDNAMES = [
    "pulled_at", "post_id", "post_url", "post_topic", "impressions",
    "total_engagements", "engagement_rate",
]


def derive_topic(url: str) -> str:
    """Best-effort, readable guess at the post topic from its URL slug.
    This is meant as a starting point, clean it up by hand afterward."""
    slug = url.rstrip("/").split("/")[-1]
    slug = re.sub(r"^s-thiru_", "", slug)
    slug = re.sub(r"-(share|ugcPost)-\d+.*$", "", slug)
    slug = unquote(slug)
    slug = unicodedata.normalize("NFKD", slug)
    slug = slug.replace("-", " ").strip()
    return slug[:80]


def post_id(url: str) -> str:
    """Extract the stable numeric post ID from a LinkedIn post URL.
    The trailing tracking suffix (e.g. -RZeU vs -nhUU) can differ between
    exports of the *same* post, so we can't match on the full URL."""
    match = re.search(r"(?:share|ugcPost)-(\d+)", url)
    return match.group(1) if match else url


def parse_top_posts(ws):
    """Merge the two side-by-side Top Posts tables (by-engagement and
    by-impressions) into one dict keyed by stable post ID."""
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row and row[0] == "Post URL":
            header_row = i
            break
    if header_row is None:
        raise ValueError("Could not find 'Post URL' header in TOP POSTS sheet")

    posts = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # Left table: Post URL, Publish Date, Engagements
        url_l, date_l, eng_l = row[0], row[1], row[2]
        if url_l:
            pid = post_id(url_l)
            posts.setdefault(pid, {})
            posts[pid]["url"] = url_l
            posts[pid]["publish_date"] = date_l
            posts[pid]["engagements"] = eng_l

        # Right table: Post URL, Publish Date, Impressions (cols E, F, G -> idx 4,5,6)
        if len(row) >= 7:
            url_r, date_r, imp_r = row[4], row[5], row[6]
            if url_r:
                pid = post_id(url_r)
                posts.setdefault(pid, {})
                posts[pid]["url"] = posts[pid].get("url") or url_r
                posts[pid]["publish_date"] = posts[pid].get("publish_date") or date_r
                posts[pid]["impressions"] = imp_r

    return posts


def parse_followers(ws):
    """Return {date_str: new_followers} from the FOLLOWERS sheet."""
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row and row[0] == "Date":
            header_row = i
            break
    if header_row is None:
        return {}

    followers = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row and row[0]:
            followers[str(row[0])] = row[1]
    return followers


def parse_engagement(ws):
    """Return {date_str: {impressions, engagements}} from the ENGAGEMENT
    sheet. This is the account-wide daily total, across all posts."""
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row and row[0] == "Date":
            header_row = i
            break
    if header_row is None:
        return {}

    daily = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row and row[0]:
            daily[str(row[0])] = {
                "impressions": int(row[1] or 0),
                "engagements": int(row[2] or 0),
            }
    return daily


def parse_discovery(ws):
    """Return the overall range summary from the DISCOVERY sheet.
    This is a single snapshot for the whole export, not per-day."""
    values = {}
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if not row or not row[0]:
            continue
        if row[0] == "Overall Performance":
            range_str = str(row[1] or "")
            if " - " in range_str:
                values["range_start"], values["range_end"] = range_str.split(" - ", 1)
            else:
                values["range_start"], values["range_end"] = "", ""
        elif row[0] == "Impressions":
            values["total_impressions"] = row[1]
        elif row[0] == "Members reached":
            values["total_members_reached"] = row[1]
    return values


def build_rows(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    posts = parse_top_posts(wb["TOP POSTS"])
    followers = parse_followers(wb["FOLLOWERS"]) if "FOLLOWERS" in wb.sheetnames else {}

    seen_dates_for_followers = set()
    rows = []

    # Sort by publish date so "first post of the day" is deterministic
    def sort_key(item):
        return str(item[1].get("publish_date") or "")

    for pid, data in sorted(posts.items(), key=sort_key):
        url = data.get("url", "")
        date_str = str(data.get("publish_date") or "")
        impressions = int(data.get("impressions") or 0)
        engagements = int(data.get("engagements") or 0)
        engagement_rate = round(engagements / impressions, 4) if impressions else 0

        new_followers = ""
        if date_str in followers and date_str not in seen_dates_for_followers:
            new_followers = followers[date_str]
            seen_dates_for_followers.add(date_str)

        rows.append({
            "post_id": pid,
            "date": date_str,
            "post_url": url,
            "post_topic": derive_topic(url),
            "content_type": "",
            "pillar": "",
            "impressions": impressions,
            "reactions": "",
            "comments": "",
            "shares": "",
            "total_engagements": engagements,
            "engagement_rate": engagement_rate,
            "new_followers": new_followers,
        })

    return rows


def build_daily_rows(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    engagement = parse_engagement(wb["ENGAGEMENT"]) if "ENGAGEMENT" in wb.sheetnames else {}
    followers = parse_followers(wb["FOLLOWERS"]) if "FOLLOWERS" in wb.sheetnames else {}

    rows = []
    for date_str, data in engagement.items():
        impressions = data["impressions"]
        engagements = data["engagements"]
        engagement_rate = round(engagements / impressions, 4) if impressions else 0
        rows.append({
            "date": date_str,
            "impressions": impressions,
            "engagements": engagements,
            "engagement_rate": engagement_rate,
            "new_followers": followers.get(date_str, 0),
        })
    return rows


def merge_and_save_daily(new_rows):
    existing = {}
    if DAILY_PATH.exists():
        with open(DAILY_PATH, newline="", encoding="utf-8") as f:
            existing = {row["date"]: row for row in csv.DictReader(f)}

    added, updated = 0, 0
    for row in new_rows:
        if row["date"] in existing:
            updated += 1
        else:
            added += 1
        existing[row["date"]] = row

    DAILY_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(DAILY_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=DAILY_FIELDNAMES)
        writer.writeheader()

        def date_key(d):
            try:
                return datetime.strptime(d, "%m/%d/%Y")
            except ValueError:
                return datetime.min

        for row in sorted(existing.values(), key=lambda r: date_key(r["date"])):
            writer.writerow({k: row.get(k, "") for k in DAILY_FIELDNAMES})

    print(f"Daily totals: {added} new day(s) added, {updated} existing day(s) refreshed.")
    print(f"  -> {DAILY_PATH.resolve()} ({len(existing)} total days)")


def append_discovery_log(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "DISCOVERY" not in wb.sheetnames:
        return
    values = parse_discovery(wb["DISCOVERY"])
    if not values:
        return

    row = {
        "pulled_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "range_start": values.get("range_start", ""),
        "range_end": values.get("range_end", ""),
        "total_impressions": values.get("total_impressions", ""),
        "total_members_reached": values.get("total_members_reached", ""),
    }

    DISCOVERY_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    file_exists = DISCOVERY_LOG_PATH.exists()
    with open(DISCOVERY_LOG_PATH, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=DISCOVERY_FIELDNAMES)
        if not file_exists:
            writer.writeheader()
        writer.writerow(row)

    print(f"Discovery log: logged this run's range summary.")
    print(f"  -> {DISCOVERY_LOG_PATH.resolve()}")


def append_snapshots(rows, xlsx_path):
    """Append one row per post for this pull, using the export's range_end
    as the 'as of' date. Never overwrites, every pull adds new rows, this
    is what makes a per-post growth-over-time chart possible."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    discovery = parse_discovery(wb["DISCOVERY"]) if "DISCOVERY" in wb.sheetnames else {}
    pulled_at = discovery.get("range_end") or datetime.now().strftime("%m/%d/%Y")

    existing_keys = set()
    file_exists = SNAPSHOT_PATH.exists()
    if file_exists:
        with open(SNAPSHOT_PATH, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                existing_keys.add((row["pulled_at"], row["post_id"]))

    SNAPSHOT_PATH.parent.mkdir(parents=True, exist_ok=True)
    added = 0
    with open(SNAPSHOT_PATH, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=SNAPSHOT_FIELDNAMES)
        if not file_exists:
            writer.writeheader()
        for row in rows:
            pid = row["post_id"]
            key = (pulled_at, pid)
            if key in existing_keys:
                continue  # already have a snapshot for this post at this pull date
            writer.writerow({
                "pulled_at": pulled_at,
                "post_id": pid,
                "post_url": row["post_url"],
                "post_topic": row["post_topic"],
                "impressions": row["impressions"],
                "total_engagements": row["total_engagements"],
                "engagement_rate": row["engagement_rate"],
            })
            added += 1

    print(f"Post snapshots: {added} new snapshot row(s) added for pull dated {pulled_at}.")
    print(f"  -> {SNAPSHOT_PATH.resolve()}")


def load_master():
    if not MASTER_PATH.exists():
        return {}
    with open(MASTER_PATH, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return {post_id(row["post_url"]): row for row in reader}


def merge_and_save(new_rows):
    existing = load_master()
    added, updated = 0, 0

    for row in new_rows:
        pid = row.pop("post_id")
        if pid in existing:
            old = existing[pid]
            # Preserve manual edits, only refresh the metrics LinkedIn reports
            for field in ("impressions", "total_engagements", "engagement_rate", "new_followers", "date", "post_url"):
                if row[field] not in ("", None):
                    old[field] = row[field]
            # Keep manually-tagged fields as-is unless they were never set
            for field in ("post_topic", "content_type", "pillar"):
                if not old.get(field):
                    old[field] = row[field]
            updated += 1
        else:
            existing[pid] = row
            added += 1

    MASTER_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(MASTER_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        for row in sorted(existing.values(), key=lambda r: r.get("date", "")):
            writer.writerow({k: row.get(k, "") for k in FIELDNAMES})

    print(f"Done. {added} new post(s) added, {updated} existing post(s) refreshed.")
    print(f"Master dataset: {MASTER_PATH.resolve()} ({len(existing)} total posts)")


def main():
    if len(sys.argv) != 2:
        print("Usage: python3 parse_linkedin_export.py path/to/export.xlsx")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    rows = build_rows(xlsx_path)
    rows_copy = [dict(r) for r in rows]  # merge_and_save pops post_id, snapshots need it intact
    merge_and_save(rows)

    print()
    append_snapshots(rows_copy, xlsx_path)

    print()
    daily_rows = build_daily_rows(xlsx_path)
    merge_and_save_daily(daily_rows)

    print()
    append_discovery_log(xlsx_path)


if __name__ == "__main__":
    main()
